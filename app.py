import streamlit as st
import pandas as pd
import openpyxl
import io
import datetime
import re
import rules_engine

# Page Configuration
st.set_page_config(
    page_title="Taxation Coding Verifier | SIX Financial",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Styling (Slate / Glassmorphism theme)
st.markdown("""
<style>
    /* Premium Slate Theme Variables */
    :root {
        --bg-color: #0F172A;
        --sidebar-bg: #1E293B;
        --card-bg: #1E293B;
        --accent: #6366F1;
        --text-main: #F8FAFC;
        --text-muted: #94A3B8;
    }
    
    /* Font styles */
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', sans-serif;
    }
    
    /* Metric Cards */
    .metric-container {
        display: flex;
        gap: 16px;
        margin-bottom: 24px;
        flex-wrap: wrap;
    }
    
    .metric-card {
        flex: 1;
        min-width: 160px;
        background: linear-gradient(135deg, #1E293B 0%, #0F172A 100%);
        border: 1px solid #334155;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1), 0 2px 4px -2px rgb(0 0 0 / 0.1);
        transition: transform 0.2s ease, border-color 0.2s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        border-color: #475569;
    }
    
    .metric-label {
        font-size: 12px;
        color: #94A3B8;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    .metric-value {
        font-size: 28px;
        font-weight: 700;
        margin-top: 8px;
        color: #F8FAFC;
    }
    
    /* Status Badges */
    .status-badge {
        display: inline-block;
        padding: 4px 8px;
        border-radius: 6px;
        font-size: 12px;
        font-weight: 600;
        text-align: center;
    }
    .status-ok { background-color: #065F46; color: #D1FAE5; }
    .status-verify { background-color: #065F46; color: #D1FAE5; border: 1px dashed #D1FAE5; }
    .status-error { background-color: #991B1B; color: #FEE2E2; }
    .status-error-verify { background-color: #991B1B; color: #FEE2E2; border: 1px dashed #FEE2E2; }
    .status-manual { background-color: #92400E; color: #FEF3C7; }
    .status-pending { background-color: #374151; color: #F3F4F6; }

    /* Custom Header Info */
    .app-header {
        background: linear-gradient(90deg, #6366F1 0%, #4F46E5 100%);
        padding: 24px;
        border-radius: 12px;
        color: white;
        margin-bottom: 24px;
        box-shadow: 0 4px 15px rgba(99, 102, 241, 0.2);
    }
</style>
""", unsafe_allow_html=True)

# App Title & Header
st.markdown("""
<div class="app-header">
    <h1 style="margin: 0; font-size: 28px; font-weight: 700; color: white;">Taxation Coding Verifier</h1>
    <p style="margin: 5px 0 0 0; font-size: 14px; opacity: 0.9; color: white;">
        Reference Data Verification Tool for SIX Financial. Upload a daily-exported Excel sheet to verify Valor codings.
    </p>
</div>
""", unsafe_allow_html=True)

# Main UI container
st.sidebar.markdown("### Upload Data")
uploaded_file = st.sidebar.file_uploader(
    "Choose GKTK Export Workbook (.xlsx)",
    type=["xlsx"],
    help="Upload the exported workbook containing 'Institution basic data' and 'Instrument basic data (Bonds)' sheets."
)

verify_clicked = st.sidebar.button(
    "Verify Coding",
    type="primary",
    disabled=(uploaded_file is None),
    use_container_width=True
)

def build_excel_report(results, counts):
    """Generates the styled Excel report for download."""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    ws_summary.append(["Status", "Count"])
    for status in ["OK", "OK (VERIFY)", "ERROR", "ERROR (VERIFY)", "PENDING", "MANUAL REVIEW"]:
        ws_summary.append([status, counts.get(status, 0)])
        
    header_fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    white_bold_font = openpyxl.styles.Font(name="Arial", size=11, bold=True, color="FFFFFF")
    arial_font = openpyxl.styles.Font(name="Arial", size=11)
    
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = white_bold_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="left")
        
    for r_idx in range(2, ws_summary.max_row + 1):
        for cell in ws_summary[r_idx]:
            cell.font = arial_font
            
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # 2. Details Sheet
    ws_details = wb.create_sheet(title="Verification Results")
    ws_details.views.sheetView[0].showGridLines = True
    
    headers = [
        "Item Key",
        "Row (Bonds sheet)",
        "Product (Prefix)",
        "Country",
        "Status",
        "Reason",
        "Expected Tax & Reporting",
        "Expected Income Code",
        "Actual Tax & Reporting",
        "Actual Income Code"
    ]
    ws_details.append(headers)
    
    status_colors = {
        "OK": openpyxl.styles.PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "OK (VERIFY)": openpyxl.styles.PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "ERROR": openpyxl.styles.PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "ERROR (VERIFY)": openpyxl.styles.PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "MANUAL REVIEW": openpyxl.styles.PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "PENDING": openpyxl.styles.PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }
    
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin', color='D1D5DB'),
        right=openpyxl.styles.Side(style='thin', color='D1D5DB'),
        top=openpyxl.styles.Side(style='thin', color='D1D5DB'),
        bottom=openpyxl.styles.Side(style='thin', color='D1D5DB')
    )
    
    for item in results:
        ws_details.append([
            item.get("Item Key", ""),
            item.get("Row (Bonds sheet)", ""),
            item.get("Product (Prefix)", ""),
            item.get("Country", ""),
            item.get("Status", ""),
            item.get("Reason", ""),
            item.get("Expected Tax & Reporting", ""),
            item.get("Expected Income Code", ""),
            item.get("Actual Tax & Reporting", ""),
            item.get("Actual Income Code", "")
        ])
        
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.font = white_bold_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
        
    ws_details.freeze_panes = "A2"
    
    for r_idx in range(2, ws_details.max_row + 1):
        status_val = ws_details.cell(row=r_idx, column=5).value
        fill_to_apply = status_colors.get(status_val)
        
        for c_idx in range(1, 11):
            cell = ws_details.cell(row=r_idx, column=c_idx)
            cell.font = arial_font
            cell.border = thin_border
            
            if c_idx in (2, 5):
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            elif c_idx == 6:
                cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
                
            if c_idx == 5 and fill_to_apply:
                cell.fill = fill_to_apply
                
    for col in ws_details.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        if col[0].column == 6:
            ws_details.column_dimensions[col_letter].width = 55
        else:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws_details.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output)
    return output.getvalue()

# Verify Action
if verify_clicked:
    st.session_state.results = None
    st.session_state.counts = None
    st.session_state.export_timestamp = None
    
    with st.spinner("Processing file and running verification engine..."):
        try:
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        except Exception as e:
            st.error(f"Failed to read the Excel file: {e}")
            st.stop()
            
        # Sheet existence check
        if 'Institution basic data' not in wb.sheetnames:
            st.error("Could not find sheet 'Institution basic data' in the uploaded workbook.")
            st.stop()
        if 'Instrument basic data (Bonds)' not in wb.sheetnames:
            st.error("Could not find sheet 'Instrument basic data (Bonds)' in the uploaded workbook.")
            st.stop()
            
        # Retrieve sheets
        inst_sheet = wb['Institution basic data']
        bonds_sheet = wb['Instrument basic data (Bonds)']
        
        # Parse institutions and bonds with column header resolution
        try:
            institutions = rules_engine.parse_institution_sheet(inst_sheet)
        except ValueError as e:
            st.error(f"Error in 'Institution basic data' sheet: {e}")
            st.stop()
            
        try:
            bonds = rules_engine.parse_bonds_sheet(bonds_sheet)
        except ValueError as e:
            st.error(f"Error in 'Instrument basic data (Bonds)' sheet: {e}")
            st.stop()
            
        # Extract File Export Timestamp
        export_date = rules_engine.extract_file_timestamp(bonds_sheet)
        st.session_state.export_timestamp = export_date
        
        # Run verification engine
        raw_results = []
        counts = {
            "OK": 0, "OK (VERIFY)": 0, "ERROR": 0,
            "ERROR (VERIFY)": 0, "PENDING": 0, "MANUAL REVIEW": 0
        }
        
        for item_key, bond in bonds.items():
            # Get institution Country for report display
            inst = institutions.get(item_key)
            country_label = inst['country_code'] if inst else "Unknown"
            
            # Execute rule decision tree
            verif = rules_engine.verify_bond(bond, institutions, export_date)
            status = verif['status']
            counts[status] += 1
            
            raw_results.append({
                "Item Key": bond['item_key'],
                "Row (Bonds sheet)": bond['row_num'],
                "Product (Prefix)": bond['prefix'],
                "Country": country_label,
                "Status": status,
                "Reason": verif['reason'],
                "Expected Tax & Reporting": verif['expected_tax'],
                "Expected Income Code": verif['expected_income'],
                "Actual Tax & Reporting": bond['taxation'] or "Blank",
                "Actual Income Code": bond['income_code'] or "Blank"
            })
            
        # Sort results by Row (Bonds sheet)
        raw_results.sort(key=lambda x: x["Row (Bonds sheet)"])
        
        # Save to Session State
        st.session_state.results = raw_results
        st.session_state.counts = counts
        st.success("Verification completed successfully!")

# Display dashboard if results exist
if "results" in st.session_state and st.session_state.results is not None:
    results = st.session_state.results
    counts = st.session_state.counts
    export_timestamp = st.session_state.export_timestamp
    
    # 1. Metric Counts Strip
    total_securities = len(results)
    errors_count = counts["ERROR"] + counts["ERROR (VERIFY)"]
    verify_count = counts["OK (VERIFY)"]
    manual_count = counts["MANUAL REVIEW"]
    pending_count = counts["PENDING"]
    ok_count = counts["OK"]
    
    st.markdown(f"""
    <div class="metric-container">
        <div class="metric-card" style="border-left: 4px solid #6366F1;">
            <div class="metric-label">Total Bonds</div>
            <div class="metric-value">{total_securities}</div>
        </div>
        <div class="metric-card" style="border-left: 4px solid #10B981;">
            <div class="metric-label">OK</div>
            <div class="metric-value">{ok_count}</div>
        </div>
        <div class="metric-card" style="border-left: 4px dashed #10B981;">
            <div class="metric-label">OK (Verify)</div>
            <div class="metric-value">{verify_count}</div>
        </div>
        <div class="metric-card" style="border-left: 4px solid #EF4444;">
            <div class="metric-label">Errors</div>
            <div class="metric-value" style="color: {'#F8FAFC' if errors_count == 0 else '#EF4444'};">{errors_count}</div>
        </div>
        <div class="metric-card" style="border-left: 4px solid #F59E0B;">
            <div class="metric-label">Manual Review</div>
            <div class="metric-value" style="color: {'#F8FAFC' if manual_count == 0 else '#F59E0B'};">{manual_count}</div>
        </div>
        <div class="metric-card" style="border-left: 4px solid #6B7280;">
            <div class="metric-label">Pending</div>
            <div class="metric-value">{pending_count}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.info(f"📅 **Export Date (File Timestamp)**: {export_timestamp.strftime('%B %d, %Y') if export_timestamp else 'N/A'}")
    
    # 2. Filters & Searches
    st.markdown("### Filter Results")
    col_search, col_filter = st.columns([1, 2])
    
    with col_search:
        search_query = st.text_input("🔍 Search by Item Key or Product Prefix", "", help="Type to filter row matches instantly.")
        
    with col_filter:
        filter_status = st.radio(
            "Status Filter",
            options=["Actionable", "All", "Errors Only", "Manual Review Only", "Pending Only", "OK Only"],
            horizontal=True
        )
        
    # Apply Filtering
    filtered_results = []
    for r in results:
        # Search match
        search_match = (
            search_query.lower() in r["Item Key"].lower() or 
            search_query.lower() in r["Product (Prefix)"].lower()
        )
        if not search_match:
            continue
            
        # Status match
        status = r["Status"]
        if filter_status == "Actionable":
            status_match = status in ["ERROR", "ERROR (VERIFY)", "MANUAL REVIEW", "PENDING"]
        elif filter_status == "Errors Only":
            status_match = status in ["ERROR", "ERROR (VERIFY)"]
        elif filter_status == "Manual Review Only":
            status_match = status == "MANUAL REVIEW"
        elif filter_status == "Pending Only":
            status_match = status == "PENDING"
        elif filter_status == "OK Only":
            status_match = status in ["OK", "OK (VERIFY)"]
        else: # "All"
            status_match = True
            
        if status_match:
            filtered_results.append(r)
            
    # 3. Download Buttons
    col_dl_filtered, col_dl_all = st.columns([1, 1])
    
    with col_dl_filtered:
        filtered_excel_data = build_excel_report(filtered_results, counts)
        st.download_button(
            label="📥 Download Filtered Report (.xlsx)",
            data=filtered_excel_data,
            file_name=f"verification_report_filtered_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    with col_dl_all:
        all_excel_data = build_excel_report(results, counts)
        st.download_button(
            label="📥 Download Complete Report (.xlsx)",
            data=all_excel_data,
            file_name=f"verification_report_all_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    # 4. Results Grid Table
    st.markdown(f"**Showing {len(filtered_results)} of {total_securities} records:**")
    
    # Custom HTML Table for High-Fidelity Badge Display
    table_rows = ""
    for r in filtered_results:
        status = r["Status"]
        if status == "OK":
            badge_class = "status-badge status-ok"
        elif status == "OK (VERIFY)":
            badge_class = "status-badge status-verify"
        elif status == "ERROR":
            badge_class = "status-badge status-error"
        elif status == "ERROR (VERIFY)":
            badge_class = "status-badge status-error-verify"
        elif status == "MANUAL REVIEW":
            badge_class = "status-badge status-manual"
        else:
            badge_class = "status-badge status-pending"
            
        table_rows += f"""
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #334155; font-weight: 600;">{r["Item Key"]}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155; text-align: center;">{r["Row (Bonds sheet)"]}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155;">{r["Product (Prefix)"]}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155;">{r["Country"]}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155; text-align: center;"><span class="{badge_class}">{status}</span></td>
            <td style="padding: 10px; border-bottom: 1px solid #334155; color: #E2E8F0;">{r["Reason"] or "—"}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155; font-size: 13px; color: #94A3B8;">{r["Expected Tax & Reporting"]}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155; font-size: 13px; color: #94A3B8;">{r["Expected Income Code"]}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155; font-size: 13px; color: #CBD5E1;">{r["Actual Tax & Reporting"]}</td>
            <td style="padding: 10px; border-bottom: 1px solid #334155; font-size: 13px; color: #CBD5E1;">{r["Actual Income Code"]}</td>
        </tr>
        """
        
    st.markdown(f"""
    <div style="overflow-x: auto; border: 1px solid #334155; border-radius: 12px; background-color: #1E293B; margin-top: 15px;">
        <table style="width: 100%; border-collapse: collapse; text-align: left; font-size: 14px;">
            <thead>
                <tr style="background-color: #0F172A; border-bottom: 2px solid #334155;">
                    <th style="padding: 12px 10px; color: #F8FAFC;">Item Key</th>
                    <th style="padding: 12px 10px; color: #F8FAFC; text-align: center;">Row</th>
                    <th style="padding: 12px 10px; color: #F8FAFC;">Product (Prefix)</th>
                    <th style="padding: 12px 10px; color: #F8FAFC;">Country</th>
                    <th style="padding: 12px 10px; color: #F8FAFC; text-align: center;">Status</th>
                    <th style="padding: 12px 10px; color: #F8FAFC; width: 250px;">Reason</th>
                    <th style="padding: 12px 10px; color: #94A3B8; font-size: 13px;">Expected Tax</th>
                    <th style="padding: 12px 10px; color: #94A3B8; font-size: 13px;">Expected Income</th>
                    <th style="padding: 12px 10px; color: #CBD5E1; font-size: 13px;">Actual Tax</th>
                    <th style="padding: 12px 10px; color: #CBD5E1; font-size: 13px;">Actual Income</th>
                </tr>
            </thead>
            <tbody>
                {table_rows if table_rows else '<tr><td colspan="10" style="padding: 20px; text-align: center; color: #94A3B8;">No records matched active search/filters.</td></tr>'}
            </tbody>
        </table>
    </div>
    """, unsafe_allow_html=True)
else:
    # Upload instructions block
    st.info("👈 Please upload a GKTK workbook in the sidebar and click **Verify Coding** to generate the verification report.")
    
    st.markdown("""
    ### Quick Instructions
    1. **Format**: The workbook must be an Excel spreadsheet (`.xlsx`) containing at least two sheets:
       - `Institution basic data` (with fields: `Item Key`, `Sector Code TK`, and `Country Code`)
       - `Instrument basic data (Bonds)` (with fields: `Item Key`, `Prefix`, `Suffix`, `Maturity`, `Interest period/usufruct start`, `Taxation and reporting`, and `Income code`)
    2. **Resilience**: The columns and headers are resolved dynamically on Row 7. Metadata rows 1–6 are safely ignored.
    3. **Status Classification**:
       - **OK**: Verified matching codings.
       - **OK (VERIFY)**: Matching coding, but verified using approximated maturity start dates (missing date verification is recommended).
       - **ERROR**: Coding mismatch error.
       - **ERROR (VERIFY)**: Coding mismatch based on approximated dates.
       - **PENDING**: Coded fields are blank but expecting values (not yet edited).
       - **MANUAL REVIEW**: Security with unresolvable/missing details (e.g. missing linked institution or maturity date).
    """)
