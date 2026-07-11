import re
import datetime
import openpyxl

def extract_code(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    if '-' in val_str:
        return val_str.split('-')[0].strip()
    return val_str

def parse_date(val):
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                pass
    return None

def extract_rate_and_is_zero(prefix):
    if not prefix:
        return False
    # Check for parenthesized portion containing "min" (case-insensitive)
    parentheses_contents = re.findall(r'\(([^)]*)\)', prefix)
    has_min = False
    for content in parentheses_contents:
        if 'min' in content.lower():
            has_min = True
            break
    if has_min:
        return False  # Treat as interest-bearing (non-zero)
        
    match = re.search(r'(\d+(?:\.\d+)?)\s*%', prefix)
    if match:
        rate = float(match.group(1))
        return rate == 0.0
    return False  # Default to interest-bearing (non-zero)

def is_muni_taxable(suffix):
    if not suffix:
        return False
    suffix_lower = suffix.lower()
    if re.search(r'\bnon[- ]?taxable\b', suffix_lower) or re.search(r'\bnontaxable\b', suffix_lower):
        return False
    elif re.search(r'\btaxable\b', suffix_lower):
        return True
    return False

def is_mortgage_related(prefix):
    if not prefix:
        return False
    prefix_lower = prefix.lower()
    keywords = [
        "mortgage",
        "morgtage",
        "remic",
        "real estate mortgage investment conduit",
        "collateralized mortgage obligation"
    ]
    return any(kw in prefix_lower for kw in keywords)

def extract_file_timestamp(sheet):
    # Scan the first 10 rows for a cell containing "Timestamp"
    for row in sheet.iter_rows(max_row=10, max_col=5, values_only=True):
        if row and len(row) > 1:
            if row[0] == "Timestamp" and row[1] is not None:
                parsed_dt = parse_date(row[1])
                if parsed_dt:
                    return parsed_dt
    return datetime.date.today()

def find_header_row_and_column_indices(sheet, required_headers):
    header_row_idx = None
    headers = []
    
    # Scan first 20 rows for a row that has "Item Key"
    for r_idx in range(1, 21):
        row_vals = [cell.value for cell in sheet[r_idx]]
        if "Item Key" in row_vals:
            header_row_idx = r_idx
            headers = [str(val).strip() if val is not None else None for val in row_vals]
            break
            
    if header_row_idx is None:
        raise ValueError("Could not find header row containing 'Item Key' in the sheet.")
        
    mapping = {}
    missing = []
    for h in required_headers:
        if h in headers:
            mapping[h] = headers.index(h)
        else:
            missing.append(h)
            
    if missing:
        raise ValueError(f"Required header(s) not found in row {header_row_idx}: {', '.join(missing)}")
        
    return header_row_idx, mapping

def parse_institution_sheet(sheet):
    required = ["Item Key", "Sector Code TK", "Country Code"]
    header_row_idx, mapping = find_header_row_and_column_indices(sheet, required)
    
    item_key_idx = mapping["Item Key"]
    sector_idx = mapping["Sector Code TK"]
    country_idx = mapping["Country Code"]
    
    institutions = {}
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        item_key = row[item_key_idx]
        if not item_key:
            continue
        item_key_str = str(item_key).strip()
        if not item_key_str or item_key_str.lower() == 'item key':
            continue
            
        if item_key_str not in institutions:
            institutions[item_key_str] = {'sector_code': None, 'country_code': None}
            
        sector = row[sector_idx]
        country = row[country_idx]
        
        if sector is not None and institutions[item_key_str]['sector_code'] is None:
            institutions[item_key_str]['sector_code'] = str(sector).strip()
        if country is not None and institutions[item_key_str]['country_code'] is None:
            institutions[item_key_str]['country_code'] = str(country).strip()
            
    return institutions

def parse_bonds_sheet(sheet):
    required = [
        "Item Key", "Prefix", "Suffix", "Maturity", 
        "Interest period/usufruct start", "Taxation and reporting", "Income code"
    ]
    header_row_idx, mapping = find_header_row_and_column_indices(sheet, required)
    
    item_key_idx = mapping["Item Key"]
    prefix_idx = mapping["Prefix"]
    suffix_idx = mapping["Suffix"]
    maturity_idx = mapping["Maturity"]
    interest_start_idx = mapping["Interest period/usufruct start"]
    taxation_idx = mapping["Taxation and reporting"]
    income_code_idx = mapping["Income code"]
    
    bonds = {}
    for r_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        item_key = row[item_key_idx]
        if not item_key:
            continue
        item_key_str = str(item_key).strip()
        if not item_key_str or item_key_str.lower() == 'item key':
            continue
            
        prefix = row[prefix_idx]
        # Skip Repeating substructures: first row must carry Prefix
        if prefix is None:
            continue
            
        if item_key_str in bonds:
            continue
            
        suffix = row[suffix_idx]
        maturity = row[maturity_idx]
        interest_start = row[interest_start_idx]
        taxation = row[taxation_idx]
        income_code = row[income_code_idx]
        
        bonds[item_key_str] = {
            'item_key': item_key_str,
            'row_num': r_idx,
            'prefix': str(prefix).strip(),
            'suffix': str(suffix).strip() if suffix is not None else None,
            'maturity': maturity,
            'interest_start': interest_start,
            'taxation': str(taxation).strip() if taxation is not None else None,
            'income_code': str(income_code).strip() if income_code is not None else None
        }
        
    return bonds

def verify_bond(bond, institutions, today_date):
    inst = institutions.get(bond['item_key'])
    if not inst:
        return {
            'status': 'MANUAL REVIEW',
            'reason': "Item Key not found in Institution basic data — cannot verify country/sector.",
            'expected_tax': 'N/A',
            'expected_income': 'N/A'
        }

    country_full = inst['country_code']
    country_code = extract_code(country_full)
    sector_full = inst['sector_code']
    sector_code = extract_code(sector_full)

    actual_tax_full = bond['taxation']
    actual_tax = extract_code(actual_tax_full)
    actual_income_full = bond['income_code']
    actual_income = extract_code(actual_income_full)

    # Step 0 — US-domicile gate
    if country_code != "333":
        expected_tax = ""
        expected_income = ""
        if actual_tax != "" or actual_income != "":
            return {
                'status': 'ERROR',
                'reason': f"Taxation Coded for Non US Domicile (Country: {country_full or 'Unknown'})",
                'expected_tax': 'Blank',
                'expected_income': 'Blank'
            }
        else:
            return {
                'status': 'OK',
                'reason': '',
                'expected_tax': 'Blank',
                'expected_income': 'Blank'
            }

    # Step 1 — Compute the duration
    maturity_date = parse_date(bond['maturity'])
    start_date = parse_date(bond['interest_start'])

    duration_approximated = False
    duration_days = None

    if not maturity_date:
        return {
            'status': 'MANUAL REVIEW',
            'reason': "Maturity date missing — cannot determine duration.",
            'expected_tax': 'N/A',
            'expected_income': 'N/A'
        }

    if not start_date:
        # Approximate duration as Maturity − export date
        approx_duration = (maturity_date - today_date).days
        if approx_duration > 365:
            duration_days = approx_duration
            duration_approximated = True
        else:
            return {
                'status': 'MANUAL REVIEW',
                'reason': "Interest period/usufruct start missing and maturity <1yr out — duration uncertain, verify manually.",
                'expected_tax': 'N/A',
                'expected_income': 'N/A'
            }
    else:
        duration_days = (maturity_date - start_date).days

    # Step 2 — Duration ≤ 183 days (any product)
    if duration_days <= 183:
        expected_tax = "2"
        expected_income = ""
        expected_tax_label = "2-US-IRS non taxable,form 1099,QI report.amount"
        expected_income_label = "Blank"
        
        is_match = (actual_tax == expected_tax) and (actual_income == expected_income)
        if is_match:
            status = "OK (VERIFY)" if duration_approximated else "OK"
            reason = "Interest period/usufruct start missing — duration approximated from today's date; please confirm manually." if duration_approximated else ""
        else:
            if actual_tax == "" and actual_income == "":
                status = "PENDING"
                reason = "US security (duration <= 183 days) — not yet coded."
            else:
                status = "ERROR (VERIFY)" if duration_approximated else "ERROR"
                reason = (
                    f"Mismatch (approximated duration <= 183 days) — expected 2/blank, got {actual_tax or 'blank'}/{actual_income or 'blank'}"
                    if duration_approximated else
                    f"Mismatch — rule applied: [duration <= 183 days]. Expected 2/blank, got {actual_tax or 'blank'}/{actual_income or 'blank'}."
                )
        return {
            'status': status,
            'reason': reason,
            'expected_tax': expected_tax_label,
            'expected_income': expected_income_label
        }

    # Step 3 — Duration > 183 days: check for a municipal issuer (Sectors 11 or 3)
    is_muni = (sector_code in ("11", "3"))

    if is_muni:
        suffix = bond['suffix']
        suffix_taxable = is_muni_taxable(suffix)
        prefix = bond['prefix']
        rate_is_zero = extract_rate_and_is_zero(prefix)

        if suffix_taxable:
            if rate_is_zero:
                expected_tax = "1"
                expected_income = "30"
                expected_tax_label = "1-US-IRS taxable and reportable w.1042S or 1099"
                expected_income_label = "30-(IRS 30) Inc.on original issue discount (OID)"
                rule_name = "Municipal, taxable (Suffix contains 'Taxable', 0% rate)"
            else:
                expected_tax = "1"
                expected_income = "1"
                expected_tax_label = "1-US-IRS taxable and reportable w.1042S or 1099"
                expected_income_label = "1-(IRS 01) Interest by US obligors gen."
                rule_name = "Municipal, taxable (Suffix contains 'Taxable', pays interest)"
        else:
            if rate_is_zero:
                expected_tax = ""
                expected_income = ""
                expected_tax_label = "Blank"
                expected_income_label = "Blank"
                rule_name = "Municipal, non-taxable (default/explicit non-taxable, 0% rate)"
            else:
                expected_tax = "2"
                expected_income = ""
                expected_tax_label = "2-US-IRS non taxable,form 1099,QI report.amount"
                expected_income_label = "Blank"
                rule_name = "Municipal, non-taxable (default, Suffix has no 'Taxable' marker) — expected 2/blank." if not suffix else "Municipal, non-taxable (explicit, Suffix contains 'Non-Taxable') — expected 2/blank."

        is_match = (actual_tax == expected_tax) and (actual_income == expected_income)

        if is_match:
            status = "OK (VERIFY)" if duration_approximated else "OK"
            reason = "Interest period/usufruct start missing — duration approximated from today's date; please confirm manually." if duration_approximated else ""
        else:
            if actual_tax == "" and actual_income == "":
                status = "PENDING"
                reason = f"US security ({rule_name}) — not yet coded."
            else:
                status = "ERROR (VERIFY)" if duration_approximated else "ERROR"
                reason = (
                    f"Mismatch (approximated muni duration >183 days) — {rule_name.split('—')[0].strip()}. Expected {expected_tax or 'blank'}/{expected_income or 'blank'}, got {actual_tax or 'blank'}/{actual_income or 'blank'}"
                    if duration_approximated else
                    f"Mismatch — rule applied: [{rule_name.split('—')[0].strip()}]. Expected {expected_tax or 'blank'}/{expected_income or 'blank'}, got {actual_tax or 'blank'}/{actual_income or 'blank'}."
                )
                if not suffix_taxable and not rate_is_zero:
                    suffix_lower = suffix.lower() if suffix else ""
                    has_explicit_non = re.search(r'\bnon[- ]?taxable\b', suffix_lower)
                    if not has_explicit_non:
                        reason = f"Municipal, non-taxable (default, Suffix has no 'Taxable' marker) — expected 2/blank."
                        if duration_approximated:
                            reason = "Interest period/usufruct start missing — duration approximated from today's date; please confirm manually. " + reason
        return {
            'status': status,
            'reason': reason,
            'expected_tax': expected_tax_label,
            'expected_income': expected_income_label
        }

    # Step 4 — Non-municipal, duration > 183 days: check the rate
    prefix = bond['prefix']
    rate_is_zero = extract_rate_and_is_zero(prefix)

    if rate_is_zero:
        expected_tax = "1"
        expected_income = "30"
        expected_tax_label = "1-US-IRS taxable and reportable w.1042S or 1099"
        expected_income_label = "30-(IRS 30) Inc.on original issue discount (OID)"
        rule_desc = "0% interest, duration >183 days"
    else:
        mortgage_related = is_mortgage_related(prefix)
        if mortgage_related:
            expected_tax = "1"
            expected_income = "2"
            expected_tax_label = "1-US-IRS taxable and reportable w.1042S or 1099"
            expected_income_label = "2-(IRS 02) Interest on real property mtg."
            rule_desc = "mortgage-related, duration >183 days"
        else:
            expected_tax = "1"
            expected_income = "1"
            expected_tax_label = "1-US-IRS taxable and reportable w.1042S or 1099"
            expected_income_label = "1-(IRS 01) Interest by US obligors gen."
            rule_desc = "interest-bearing, non-mortgage, duration >183 days"

    is_match = (actual_tax == expected_tax) and (actual_income == expected_income)

    if is_match:
        status = "OK (VERIFY)" if duration_approximated else "OK"
        reason = "Interest period/usufruct start missing — duration approximated from today's date; please confirm manually." if duration_approximated else ""
    else:
        if actual_tax == "" and actual_income == "":
            status = "PENDING"
            reason = f"US security ({rule_desc}) — not yet coded."
        else:
            status = "ERROR (VERIFY)" if duration_approximated else "ERROR"
            reason = (
                f"Mismatch (approximated duration >183 days) — rule applied: [{rule_desc}]. Expected {expected_tax or 'blank'}/{expected_income or 'blank'}, got {actual_tax or 'blank'}/{actual_income or 'blank'}."
                if duration_approximated else
                f"Mismatch — rule applied: [{rule_desc}]. Expected {expected_tax or 'blank'}/{expected_income or 'blank'}, got {actual_tax or 'blank'}/{actual_income or 'blank'}."
            )
            if rate_is_zero:
                reason = f"Mismatch — rule applied: [0% interest, duration >183 days]. Expected {expected_tax}/{expected_income}, got {actual_tax or 'blank'}/{actual_income or 'blank'}."
                if duration_approximated:
                    reason = "Interest period/usufruct start missing — duration approximated from today's date; please confirm manually. " + reason

    return {
        'status': status,
        'reason': reason,
        'expected_tax': expected_tax_label,
        'expected_income': expected_income_label
    }
